import getopt
import sys
from tqdm import tqdm
from openpyxl import load_workbook
from pygoogletranslation import Translator

def Google_Fanyi(file_home, sheet):
    try:
        wb = load_workbook(filename=file_home)  # 打开excel文件
    except:
        print("错误：打开'" + file_home + "'的文件失败，请检查文件路径是否有误！！！")
        sys.exit()
    try:
        ws = wb[sheet]  # 根据Sheet1这个sheet名字来获取该sheet
    except:
        print("错误：打开'" + sheet + "'表格失败，请检查表名是否有误！！！")
        sys.exit()
    print("准备开始翻译的Excel文件路径为:" + file_home +"，表名为:" + sheet)
    # 更换谷歌翻译地址
    translator = Translator(service_url='translate.google.cn')

    nrows = ws.max_row
    ncols = ws.max_column
    print("行数:%d\t列数:%d"%(nrows,ncols))

    input("请确定是否关闭Excel文件（输入任意键开始翻译）")
    pbar = tqdm(total=nrows * ncols)

    for row in range(1, nrows + 1):
        for col in range(1, ncols + 1):
            pbar.update(1)
            if ws.cell(row, col).value != None:
                ws.cell(row, col).value = translator.translate(ws.cell(row, col).value, dest='zh-CN').text

    try:
        wb.save(file_home)  # 保存修改后的excel
    except:
        print("错误：未关闭Excel表格文件")
        sys.exit()



if __name__ == '__main__':
    opts, args = getopt.getopt(sys.argv[1:], "hf:s:", ["file=", "sheet="])
    if len(opts) != 0:
        for opt, arg in opts:
            if opt == '-h':
                print("使用格式为：\npython Excel_Google.py -f <Execl文件位置> -s <表名>\n"
                    + "PS:如果第一次使用请安装以下包：openpyxl、pygoogletranslation。\n"
                    + "   并修改pygoogletranslation源码中的utils.py的第8行为：\n"
                    + "       from pygoogletranslation.models import TranslatedPart\n"
                    + "#开始运行代码时不可打开Excel文件#")
                sys.exit()
            elif opt in ("-f", "--file"):
                file_home = arg
            elif opt in ("-s", "--sheet"):
                sheet = arg
            else:
                print("错误：参数有误，python Excel_Google.py -h获取详情")
                sys.exit()
        try:
            Google_Fanyi(file_home, sheet)
        except:
            print("警告：参数不完整或发生错误，使用python Excel_Google.py -h获取详情")
            sys.exit()
        print("Enjoy it!")
    else:
        print("错误：参数为空，使用python Excel_Google.py -h获取详情\n"
            + "PS:如果第一次使用请安装以下包：openpyxl、pygoogletranslation。\n"
            + "   并修改pygoogletranslation源码中的utils.py的第8行为：\n"
            + "       from pygoogletranslation.models import TranslatedPart\n"
            + "#开始运行代码时不可打开Excel文件#")
        sys.exit()
    