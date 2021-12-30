import getopt
import sys
import os.path
from tqdm import tqdm
from openpyxl import load_workbook
from pygoogletranslation import Translator

SUPPORTED_EXCLE = ('.xlsx', '.xlsm', '.xltx', '.xltm')

def Excel_Translate(file_home):
	try:
		wb = load_workbook(filename=file_home)  # 打开excel文件
	except:
		print("错误：打开'" + file_home + "'的文件失败，请检查文件路径是否有误！！！")
		sys.exit()
	try:
		sheet = input("请输入需要翻译的Sheet：")
		ws = wb[sheet]  # 根据sheet名字来获取该sheet
	except:
		print("错误：打开'" + sheet + "'表格失败，请检查表名是否有误！！！")
		sys.exit()
	print("准备开始翻译的Excel文件路径为:" + file_home +"，表名为:" + sheet)
	# 更换谷歌翻译地址
	translator = Translator(service_url='translate.google.cn')

	nrows = ws.max_row
	ncols = ws.max_column
	print("行数:%d\t列数:%d"%(nrows,ncols))

	input("请确定是否关闭Excel文件（输入任意键开始翻译）")
	pbar = tqdm(total=nrows * ncols)

	for row in range(1, nrows + 1):
		for col in range(1, ncols + 1):
			pbar.update(1)
			if ws.cell(row, col).value != None:
				ws.cell(row, col).value = translator.translate(ws.cell(row, col).value, dest='zh-CN').text

	try:
		wb.save(file_home)  # 保存修改后的excel
	except:
		print("错误：未关闭Excel表格文件")
		sys.exit()



if __name__ == '__main__':
	opts, args = getopt.getopt(sys.argv[1:], "hf:", ["file="])
	if len(opts) != 0:
		for opt, arg in opts:
			if opt == '-h':
				print("使用格式为：\npython Google_Translate_Document.py -f <文档文件位置>\n"
					+ "现仅支持后缀为xlsx、xlsm、xltx、xltm的文档\n"
					+ "PS:如果第一次使用请安装以下包：openpyxl、pygoogletranslation和tqdm。\n"
					+ "   并修改pygoogletranslation源码中的utils.py的第8行为：\n"
					+ "       from pygoogletranslation.models import TranslatedPart\n"
					+ "#开始运行代码时不可打开文档文件#")
				sys.exit()
			elif opt in ("-f", "--file"):
				file_home = arg
			else:
				print("错误：参数有误，python Google_Translate_Document.py -h获取详情")
				sys.exit()
		file_format = os.path.splitext(file_home)[-1].lower()
		try:
			if file_format in SUPPORTED_EXCLE:
				Excel_Translate(file_home)
			else:
				print("错误：现仅支持后缀为xlsx、xlsm、xltx、xltm的文档")
		except:
			print("警告：参数不完整或发生错误，使用python Google_Translate_Document.py -h获取详情")
			sys.exit()
		# print("Enjoy it!")
	else:
		print("错误：参数为空，使用python Google_Translate_Document.py -h获取详情\n"
			+ "现仅支持后缀为xlsx、xlsm、xltx、xltm的文档\n"
			+ "PS:如果第一次使用请安装以下包：openpyxl、pygoogletranslation和tqdm。\n"
			+ "   并修改pygoogletranslation源码中的utils.py的第8行为：\n"
			+ "       from pygoogletranslation.models import TranslatedPart\n"
			+ "#开始运行代码时不可打开文档文件#")
		sys.exit()
    